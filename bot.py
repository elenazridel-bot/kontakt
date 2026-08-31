"""
Telegram-бот, который собирает присылаемые ему контакты и описания
"чем человек может помочь" в один Excel-файл.

Формат сообщения от пользователя:
    Имя, телефон/юзернейм
    Чем может помочь

(первая строка — контакт, остальное — описание помощи). Если сообщение
состоит из одной строки, бот пытается отделить контакт от описания по
разделителю (" - ", " — ", ":" и т.п.), а если не находит — просто
сохраняет всё как есть, ничего не теряя.

Также поддерживается пересылка/отправка контакта через встроенную кнопку
Telegram "Отправить контакт".

Команда /export в любой момент присылает текущий накопленный файл.
"""

import asyncio
import logging
import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from telegram import ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

load_dotenv()

BOT_TOKEN = os.environ.get("BOT_TOKEN")
EXCEL_FILE = os.environ.get("EXCEL_FILE", "contacts.xlsx")

HEADERS = ["Дата и время", "Добавил(а)", "Контакт", "Чем может помочь", "Исходное сообщение"]

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# Простая блокировка, чтобы параллельные сообщения не затирали друг друга
# при чтении/записи одного и того же xlsx-файла.
file_lock = asyncio.Lock()


def ensure_workbook() -> None:
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Контакты"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)


def append_row(row: list) -> None:
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(row)
    wb.save(EXCEL_FILE)


def split_contact_and_help(text: str) -> tuple[str, str]:
    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    if not lines:
        return "", ""

    if len(lines) > 1:
        return lines[0], "\n".join(lines[1:]).strip()

    single_line = lines[0]
    for sep in (" — ", " – ", " - ", ": ", ","):
        if sep in single_line:
            contact, help_text = single_line.split(sep, 1)
            return contact.strip(), help_text.strip()

    return single_line, ""


def sender_label(update: Update) -> str:
    user = update.effective_user
    if user is None:
        return ""
    if user.username:
        return f"@{user.username}"
    return user.full_name or str(user.id)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "Привет! Присылай мне контакты и то, чем человек может помочь — "
        "я соберу всё в один Excel-файл.\n\n"
        "Формат сообщения:\n"
        "Имя, телефон или юзернейм\n"
        "Чем может помочь\n\n"
        "Можно прислать это одним сообщением в две строки, а можно просто "
        "написать в одну строку — я постараюсь сам разделить контакт и "
        "описание. Также можно отправить контакт через скрепку → «Контакт».\n\n"
        "Команда /export — прислать текущий файл со всеми контактами.",
        reply_markup=ReplyKeyboardRemove(),
    )


async def export_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    async with file_lock:
        ensure_workbook()
        with open(EXCEL_FILE, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(EXCEL_FILE),
                caption="Текущий список контактов.",
            )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = update.message.text or ""
    contact, help_text = split_contact_and_help(text)

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        sender_label(update),
        contact,
        help_text,
        text,
    ]

    async with file_lock:
        await asyncio.to_thread(append_row, row)

    await update.message.reply_text("Записал ✅")


async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    contact = update.message.contact
    name = " ".join(filter(None, [contact.first_name, contact.last_name])).strip()
    phone = contact.phone_number or ""
    contact_value = ", ".join(filter(None, [name, phone]))
    help_text = update.message.caption or ""

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        sender_label(update),
        contact_value,
        help_text,
        contact_value,
    ]

    async with file_lock:
        await asyncio.to_thread(append_row, row)

    await update.message.reply_text("Контакт записал ✅")


def main() -> None:
    if not BOT_TOKEN:
        raise SystemExit(
            "Не задан BOT_TOKEN. Создайте файл .env на основе .env.example "
            "и укажите токен, полученный у @BotFather."
        )

    ensure_workbook()

    application = Application.builder().token(BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", start))
    application.add_handler(CommandHandler("export", export_file))
    application.add_handler(filters_text_handler())
    application.add_handler(MessageHandler(filters.CONTACT, handle_contact))

    logger.info("Бот запущен, файл: %s", EXCEL_FILE)
    application.run_polling(allowed_updates=Update.ALL_TYPES)


def filters_text_handler() -> MessageHandler:
    return MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)


if __name__ == "__main__":
    main()
